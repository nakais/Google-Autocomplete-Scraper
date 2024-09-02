from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime
import time

# Load the Excel file
excel_file_path = r"C:\Users\Msi\Downloads\Documents\4BeatsQ1.xlsx"
workbook = load_workbook(excel_file_path)

# Get today's day
today = datetime.datetime.today().strftime('%A')

# Select the sheet based on today's day
sheet = workbook[today]

# Set up Selenium WebDriver
driver = webdriver.Chrome()
driver.get("https://www.google.com")

def collect_autocomplete_suggestions(keyword):
    retry_attempts = 3  # Number of retries for stale element reference exception
    for attempt in range(retry_attempts):
        try:
            # Wait explicitly until the suggestions are present
            WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='wM6W7d']/span")))

            # Find the autocomplete suggestions
            suggestions = driver.find_elements(By.XPATH, "//div[@class='wM6W7d']/span")
            options = [suggestion.text.strip() for suggestion in suggestions if suggestion.text.strip()]

            return options
        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {e}")
            if attempt < retry_attempts - 1:  # Try again
                time.sleep(1)
                continue
            else:  # Raise exception after all attempts fail
                return []  # Return an empty list if all attempts fail

# Loop through the keywords and perform searches
for row in range(3, sheet.max_row + 1):
    keyword = sheet.cell(row=row, column=3).value
    if keyword:
        print(f"Collecting autocomplete suggestions for keyword: {keyword}")
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)

        # Add a delay to wait for suggestions to appear
        time.sleep(2)  # Adjust this delay if necessary

        # Collect autocomplete suggestions
        suggestions = collect_autocomplete_suggestions(keyword)

        # Find the longest and shortest suggestion based on string length
        if suggestions:
            longest_option = max(suggestions, key=len)
            shortest_option = min(suggestions, key=len)
        else:
            longest_option = "No suggestions"
            shortest_option = "No suggestions"

        # Update the Excel sheet with the longest and shortest options
        sheet.cell(row=row, column=4).value = longest_option
        sheet.cell(row=row, column=5).value = shortest_option

    else:
        print("Keyword is None, skipping...")

# Save the workbook
workbook.save(excel_file_path)

# Close the browser when done
driver.quit()
